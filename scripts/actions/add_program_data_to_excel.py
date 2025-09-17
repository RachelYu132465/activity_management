"""將 program_data.json 中的資料填入 Excel。

使用方式：

```
python -m scripts.actions.add_program_data_to_excel --input 原始檔案.xlsx --program-id 2 --output 輸出檔案.xlsx
```

會在指定的試算表中新增下列欄位（如不存在）：

* ``program_data.eventNames[0]``
* ``program_data.eventNames[1]``
* ``program_data.date``
* ``program_data.locations[0]``
* ``program_data.locations[1]``

並將對應的 program_data 資料填入每一列（只對已有資料的列進行填寫）。
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Dict, Iterable, List, Tuple

from scripts.core.bootstrap import initialize
from scripts.core.data_util import load_program_by_id


ProgramColumns = List[Tuple[str, str]]


def _require_openpyxl():
    try:
        from openpyxl import load_workbook  # type: ignore
    except ModuleNotFoundError as exc:  # pragma: no cover - optional dependency
        raise SystemExit(
            "此腳本需要 openpyxl 套件，請先安裝：pip install openpyxl"
        ) from exc
    return load_workbook


def _collect_program_columns(program: Dict[str, object]) -> ProgramColumns:
    event_names: List[str] = []
    raw_event_names = program.get("eventNames")
    if isinstance(raw_event_names, Iterable) and not isinstance(raw_event_names, (str, bytes)):
        for name in raw_event_names:
            if isinstance(name, str):
                event_names.append(name)
    elif isinstance(raw_event_names, str):
        event_names.append(raw_event_names)

    locations: List[str] = []
    raw_locations = program.get("locations")
    if isinstance(raw_locations, Iterable) and not isinstance(raw_locations, (str, bytes)):
        for loc in raw_locations:
            if isinstance(loc, str):
                locations.append(loc)
    elif isinstance(raw_locations, str):
        locations.append(raw_locations)

    raw_date = program.get("date")
    date = raw_date if isinstance(raw_date, str) else ""

    columns: ProgramColumns = [
        ("program_data.eventNames[0]", event_names[0] if len(event_names) > 0 else ""),
        ("program_data.eventNames[1]", event_names[1] if len(event_names) > 1 else ""),
        ("program_data.date", date or ""),
        ("program_data.locations[0]", locations[0] if len(locations) > 0 else ""),
        ("program_data.locations[1]", locations[1] if len(locations) > 1 else ""),
    ]
    return columns


def _ensure_headers(ws, headers: ProgramColumns) -> Dict[str, int]:
    header_rows = list(ws.iter_rows(min_row=1, max_row=1))
    header_cells = header_rows[0] if header_rows else []
    existing_headers = {cell.value: cell.col_idx for cell in header_cells if cell.value}
    header_positions: Dict[str, int] = {}

    next_col = max(existing_headers.values(), default=0)
    for header, _ in headers:
        if header in existing_headers:
            header_positions[header] = existing_headers[header]
            continue
        next_col += 1
        ws.cell(row=1, column=next_col, value=header)
        header_positions[header] = next_col
    return header_positions


def fill_program_data(
    input_path: Path,
    output_path: Path,
    program_id: str,
    sheet_name: str | None = None,
) -> None:
    load_workbook = _require_openpyxl()
    wb = load_workbook(input_path)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise SystemExit(
                "指定的工作表 '{}' 不存在，現有工作表：{}".format(sheet_name, ", ".join(wb.sheetnames))
            )
        ws = wb[sheet_name]
    else:
        ws = wb.active

    program = load_program_by_id(program_id, fallback_to_first=False)
    columns = _collect_program_columns(program)

    # 記錄原本的資料欄數，判斷哪些列需要填寫
    original_max_col = ws.max_column
    header_positions = _ensure_headers(ws, columns)

    for row_idx in range(2, ws.max_row + 1):
        row_has_value = False
        for col_idx in range(1, original_max_col + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            val = cell.value
            if val not in (None, ""):
                row_has_value = True
                break
        if not row_has_value:
            continue

        for header, value in columns:
            target_col = header_positions[header]
            ws.cell(row=row_idx, column=target_col, value=value)

    wb.save(output_path)
    wb.close()


def main() -> None:
    parser = argparse.ArgumentParser(description="將 program_data 欄位填入 Excel")
    parser.add_argument("--input", required=True, help="來源 Excel 檔案路徑")
    parser.add_argument(
        "--output",
        help="輸出 Excel 檔案路徑（預設覆寫來源檔案）",
    )
    parser.add_argument("--program-id", required=True, help="program_data.json 中的 id")
    parser.add_argument("--sheet-name", help="要處理的工作表名稱（預設為第一個工作表）")
    args = parser.parse_args()

    initialize()

    input_path = Path(args.input).expanduser().resolve()
    if not input_path.exists():
        raise SystemExit("找不到來源檔案: {}".format(input_path))

    output_path = Path(args.output).expanduser().resolve() if args.output else input_path
    output_path.parent.mkdir(parents=True, exist_ok=True)

    fill_program_data(input_path, output_path, args.program_id, args.sheet_name)
    print("已更新 Excel: {}".format(output_path))


if __name__ == "__main__":  # pragma: no cover - CLI entry point
    main()

